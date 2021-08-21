#! /usr/bin/python

import openpyxl
import requests
from datetime import datetime
from bs4 import BeautifulSoup
import pywhatkit
import time
from twilio.rest import Client

ExcelPath = r'C:\Users\Janardhan\Desktop\stock_analysis\Share_list.xlsx'
wb_obj = openpyxl.load_workbook(ExcelPath)

sheet_obj = wb_obj.active
max__rows=sheet_obj.max_row

my_wb = openpyxl.Workbook()
my_sheet = my_wb.active

while True:

  now = datetime.now()
  current_date = now.strftime("%m/%d/%y")
  current_time = now.strftime("%H:%M:%S")
  count = 0
  i = 2
  message = ["Date : "+current_date,"Time : "+current_time]
  while (i <= max__rows):
    # Getting the details from excel sheet
    Share_Name = sheet_obj.cell(row=i, column=1).value
    Share_URL = sheet_obj.cell(row=i, column=2).value
    limit_value=sheet_obj.cell(row = i, column = 3).value #to get the limit value

    # Getting the details of share from website
    response = requests.get(Share_URL, timeout=240)
    page_content = BeautifulSoup(response.content, "html.parser")

    Volume = page_content.find("span", attrs={"id": 'Volume'}).get_text()
    current_Value = page_content.find("span", attrs={"id": 'ltpid'}).get_text()
    Previous_Close = page_content.find("span", attrs={"id": 'PrevClose'}).get_text()
    high_low = page_content.find("span", attrs={"id": 'highlow'}).get_text()

    now = datetime.now()
    current_date = now.strftime("%m/%d/%y")
    current_time = now.strftime("%H:%M:%S")

    c7 = my_sheet.cell(row=1, column=1)
    c7.value = "Name"
    c8 = my_sheet.cell(row=1, column=2)
    c8.value = "Volume"
    c9 = my_sheet.cell(row=1, column=3)
    c9.value = "Previous Close"
    c13 = my_sheet.cell(row=1, column=4)
    c13.value = "High-Low"
    c10 = my_sheet.cell(row=1, column=5)
    c10.value = "Date"
    c11 = my_sheet.cell(row=1, column=6)
    c11.value = "Time"
    c12 = my_sheet.cell(row=1, column=7)
    c12.value = "Value"

    c1 = my_sheet.cell(row=i, column=1)
    c1.value = Share_Name
    c2 = my_sheet.cell(row=i, column=2)
    c2.value = Volume
    c3 = my_sheet.cell(row=i, column=3)
    c3.value = Previous_Close
    c14 = my_sheet.cell(row=i, column=4)
    c14.value = high_low
    c4 = my_sheet.cell(row=i, column=5)
    c4.value = current_date
    c5 = my_sheet.cell(row=i, column=6)
    c5.value = current_time
    c6 = my_sheet.cell(row=i, column=7)
    c6.value = current_Value


    if float(current_Value.replace(',','')) > float(Previous_Close.replace(',','')) or float(current_Value.replace(',','')) > float(limit_value):
      message.append(Share_Name + " : " + current_Value)
      count+=1

    my_wb.save(r"C:\Users\Janardhan\Desktop\stock_analysis\Data.xlsx")

    i += 1

  new_message = '\n'.join(message)
  # print(new_message)

  if count>0:

    # For whatsapp message use these
    current = datetime.now()
    current_date = current.strftime("%m/%d/%y")
    current_time = current.strftime("%H:%M:%S")
    pywhatkit.sendwhatmsg("+919505044413", new_message, current.hour, current.minute + 2)

    # For sms message use these
    #client = Client("AC3d1eb26a41cffc59532854968f078ef5", "d5a965e1aac98f4d7a53fde691207a55")
    #client.messages.create(to="+919505044413", from_="+12023355254", body=new_message)
    print("done")

  time.sleep(10)
